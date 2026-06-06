"""
job_tracker.py — Appends new job matches to job_tracker.xlsx
Run automatically by main.py after each scrape.
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TRACKER_FILE = "job_tracker.xlsx"

HEADERS = [
    "Scraped At",
    "Title",
    "Company",
    "Site",
    "Score",
    "Positives",
    "URL",
    "Description",    # ← add this
    "Applied?",
    "Applied Date",
    "Notes",
]

# Column widths
COL_WIDTHS = {
    "A": 18,   # Scraped At
    "B": 45,   # Title
    "C": 30,   # Company
    "D": 20,   # Site
    "E": 8,    # Score
    "F": 40,   # Positives
    "G": 60,   # URL
    "H": 60,   # Description  ← add this
    "I": 10,   # Applied?
    "J": 15,   # Applied Date
    "K": 30,   # Notes
}

# Header fill colours per site
SITE_COLORS = {
    "MakeItInGermany (BA API)": "D6EAF8",
    "JobTeaser (OVGU)":         "D5F5E3",
}
DEFAULT_COLOR = "F2F3F4"


def _header_fill(site: str) -> PatternFill:
    color = SITE_COLORS.get(site, DEFAULT_COLOR)
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _setup_sheet(ws) -> None:
    """Write header row with formatting."""
    header_font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Auto-filter on headers
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def clean_old_jobs(days: int = 30) -> None:
    """Remove rows older than N days from the tracker."""
    if not os.path.exists(TRACKER_FILE):
        return

    wb = load_workbook(TRACKER_FILE)
    ws = wb.active
    cutoff = datetime.now() - timedelta(days=days)

    rows_to_delete = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if not cell_value:
            continue
        try:
            row_date = datetime.strptime(str(cell_value)[:16], "%Y-%m-%d %H:%M")
            if row_date < cutoff:
                rows_to_delete.append(row)
        except Exception:
            continue

    # Delete in reverse order so row numbers don't shift
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    wb.save(TRACKER_FILE)
    if rows_to_delete:
        print(f"[tracker] Removed {len(rows_to_delete)} old job(s) from tracker")


def save_jobs_to_excel(jobs: list[dict]) -> None:
    """Append new jobs to job_tracker.xlsx, creating it if needed."""
    if not jobs:
        return

    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Load or create workbook
    if os.path.exists(TRACKER_FILE):
        wb = load_workbook(TRACKER_FILE)
        ws = wb.active
        # Check if headers exist; if not, add them (first run on old file)
        if ws.cell(row=1, column=1).value != "Scraped At":
            ws.insert_rows(1)
            _setup_sheet(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Tracker"
        _setup_sheet(ws)

    # Find next empty row
    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(row=2, column=1).value is None:
        next_row = 2  # empty sheet

    normal_font = Font(name="Arial", size=9)
    url_font    = Font(name="Arial", size=9, color="1A5276", underline="single")
    center      = Alignment(horizontal="center", vertical="center")
    wrap        = Alignment(wrap_text=True, vertical="top")

    for job in jobs:
        site      = job.get("site", "")
        title     = job.get("title", "")
        company   = job.get("company", "")
        score     = job.get("score", 0)
        positives = ", ".join(job.get("positives", []))
        url       = job.get("url", "")

        row_fill = PatternFill("solid", fgColor=SITE_COLORS.get(site, DEFAULT_COLOR))

        row_data = [
            scraped_at,
            title,
            company,
            site,
            score,
            positives,
            url,
            job.get("description", ""),    # Description
            "No",   # Applied?
            "",     # Applied Date
            "",     # Notes
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.border = _thin_border()
            cell.fill = row_fill

            if col == 7:  # URL column
                cell.font = url_font
                cell.hyperlink = url
                cell.alignment = wrap
            elif col in (2, 6, 8, 11):  # Title, Positives, Notes
                cell.font = normal_font
                cell.alignment = wrap
            elif col in (5, 9):  # Score, Applied?
                cell.font = normal_font
                cell.alignment = center
            else:
                cell.font = normal_font
                cell.alignment = Alignment(vertical="top")

        ws.row_dimensions[next_row].height = 35
        next_row += 1

    wb.save(TRACKER_FILE)
    print(f"[tracker] ✓ Added {len(jobs)} job(s) to {TRACKER_FILE}")
